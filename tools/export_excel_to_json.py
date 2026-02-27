import json
import sys
from pathlib import Path
from openpyxl import load_workbook

SHEET_CONVENTION = "Conventies"
SHEET_VERSION = "Versie toetsingsregel"

def norm(v) -> str:
    return "" if v is None else str(v).strip()

def main(xlsx_path: str, out_json_path: str) -> None:
    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        raise FileNotFoundError(f"Excel not found: {xlsx}")

    wb = load_workbook(filename=xlsx, data_only=True)

    # 1) latest rule version: sheet "Versie toetsingsregel", cell B1
    latest = ""
    if SHEET_VERSION in wb.sheetnames:
        ws_v = wb[SHEET_VERSION]
        latest = norm(ws_v.cell(row=1, column=2).value)  # B1

    # 2) conventies rows: sheet "Conventies"
    if SHEET_CONVENTION not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_CONVENTION}' not found in {xlsx.name}")

    ws = wb[SHEET_CONVENTION]

    # We export the same columns you read in Java (0-based -> 1-based):
    # Java getCell(1..8) == Excel columns 2..9
    rules = []
    for r in range(2, ws.max_row + 1):  # start at row 2 (skip header)
        iri_suffix = norm(ws.cell(row=r, column=2).value)      # col 2  (Java cell 1)
        object_id_required = norm(ws.cell(row=r, column=3).value)  # col 3 (Java cell 2) "ja/nee"
        aas_regex = norm(ws.cell(row=r, column=4).value)       # col 4  (Java cell 3)
        aas_opbouw = norm(ws.cell(row=r, column=5).value)      # col 5  (Java cell 4)
        aas_voorbeeld = norm(ws.cell(row=r, column=6).value)   # col 6  (Java cell 5)
        oms_tpl = norm(ws.cell(row=r, column=7).value)         # col 7  (Java cell 6)
        oms_uitleg = norm(ws.cell(row=r, column=8).value)      # col 8  (Java cell 7)
        oms_voorbeeld = norm(ws.cell(row=r, column=9).value)   # col 9  (Java cell 8)

        # 1-op-1 export: we keep row even if some values are empty;
        # you can choose to skip fully empty rows:
        if all(not x for x in [iri_suffix, object_id_required, aas_regex, aas_opbouw, aas_voorbeeld, oms_tpl, oms_uitleg, oms_voorbeeld]):
            continue

        rules.append({
            "iri": iri_suffix,
            "objectIdRequired": object_id_required,  # keep "ja/nee" as-is (no boolean conversion)
            "aasRegex": aas_regex,
            "aasOpbouw": aas_opbouw,
            "aasVoorbeeld": aas_voorbeeld,
            "omschrijvingTemplate": oms_tpl,
            "omschrijvingUitleg": oms_uitleg,
            "omschrijvingVoorbeeld": oms_voorbeeld,
            "row": r  # helpful for debugging; remove if you don't want it
        })

    out = {
        "latestRuleVersion": latest,
        "conventies": rules
    }

    out_path = Path(out_json_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {out_path} with {len(rules)} rows (latestRuleVersion='{latest}')")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python tools/export_excel_to_json.py <input.xlsx> <output.json>")
        sys.exit(2)
    main(sys.argv[1], sys.argv[2])
